"""
Regenerates paper_trading_performance.xlsx from the latest cleaned closed-
trade data. Rerunnable: the merge workflow refreshes paper_trades_final.csv,
then this rebuilds the workbook from it.

Layout:
  - Summary: overall KPIs, wallet vs our simulated performance, at a glance.
  - Daily Performance: one row per calendar day, with the two charts.
  - By Wallet: one row per tracked wallet.
  - Raw Trades: the underlying per-trade data everything else is built from.

Every aggregate is written as a computed VALUE, not a formula. It used to be
the other way round -- SUMIFS/COUNTIFS over Raw Trades, so the workbook would
recalculate if anyone edited that sheet by hand. That traded away far more
than it bought: openpyxl cannot write a formula's cached result, so every
aggregate cell was blank in anything that does not evaluate formulas on open
(mobile viewers, Google Sheets/Quick Look previews, pandas). The workbook
read as an empty grid everywhere except desktop Excel. And the hand-edit case
it was protecting never existed: reports/*.xlsx is gitignored and regenerated
from the CSV on every refresh, so edits to it are overwritten, not preserved.
Values it is; rerun this script to update.
"""

import datetime
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CSV_PATH = sys.argv[1] if len(sys.argv) > 1 else "paper_trades_final.csv"
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else "paper_trading_performance.xlsx"

FONT_NAME = "Arial"

HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
HEADER_FILL = "1F4E78"
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
LABEL_FONT = Font(name=FONT_NAME, bold=True)
BODY_FONT = Font(name=FONT_NAME)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9)
SMALL_NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8)

SOL_FMT = '#,##0.000;(#,##0.000);"-"'
PCT_FMT = "0.0%"
DATE_FMT = "yyyy-mm-dd"
DATETIME_FMT = "yyyy-mm-dd hh:mm"

STALE_HOURS = 12  # flag if the newest trade is older than this -- likely means the CI pipeline stopped

# Mirrors dedupe_and_clean.py; only used to recompute the flag when reading a
# CSV written before that script started emitting it.
BLOWUP_LO, BLOWUP_HI = 0.33, 3.0


def style_header_row(ws, row, n_cols):
    fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def put(ws, cell, value, font=BODY_FONT, fmt=None):
    c = ws[cell]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    return c


def main():
    df = pd.read_csv(CSV_PATH).sort_values("epoch_ms").reset_index(drop=True)

    # dedupe_and_clean.py used to drop blow-up fills outright; it now flags
    # them. Recompute when reading a CSV written before that change so this
    # script gives the same answer on any vintage of the file.
    if "blowup" not in df.columns:
        ratio = df["our_lamports_spent"] / df["wallet_lamports_spent"].replace(0, float("nan"))
        df["fill_ratio"] = ratio
        df["blowup"] = ~ratio.between(BLOWUP_LO, BLOWUP_HI)
    df["blowup"] = df["blowup"].astype(bool)

    # Collector v1 dropped reverting BUYS; v2 recorded those but still dropped
    # whole round trips whose SELL breached the bound; v3 records both. Rows
    # from an older collector genuinely carry none of the dropped cases, so
    # defaulting their flags to 0 states what those files actually contain --
    # the missing rows are simply absent and unrecoverable.
    for col, default in (("would_have_reverted", 0), ("sell_would_have_reverted", 0), ("collector_version", 1)):
        if col not in df.columns:
            df[col] = default
        df[col] = df[col].fillna(default).astype(int)

    # Raw Trades carries every row; the aggregates use the clean cohort so
    # figures stay comparable across collector versions. Reverting fills are
    # held out on the same basis as blow-ups: they are what a wider-slippage
    # bot would have eaten, not what the measured strategy did. Both cohorts
    # are sized explicitly in Summary rather than left invisible.
    reverting = (df["would_have_reverted"] == 1) | (df["sell_would_have_reverted"] == 1)
    clean = df[~df["blowup"] & ~reverting]

    n, n_clean = len(df), len(clean)
    n_blow = int(df["blowup"].sum())
    n_buy_rev = int((df["would_have_reverted"] == 1).sum())
    n_sell_rev = int((df["sell_would_have_reverted"] == 1).sum())

    newest_epoch_ms = int(df["epoch_ms"].max())
    now_ms = datetime.datetime.now(datetime.timezone.utc).timestamp() * 1000
    age_hours = (now_ms - newest_epoch_ms) / 3600000
    if age_hours > STALE_HOURS:
        print(f"WARNING: newest trade is {age_hours:.1f}h old (> {STALE_HOURS}h) -- "
              f"the data-collection pipeline may have stopped.", file=sys.stderr)

    df["date"] = pd.to_datetime(df["epoch_ms"], unit="ms").dt.date
    clean = clean.assign(date=pd.to_datetime(clean["epoch_ms"], unit="ms").dt.date)

    # Row labels come from the clean cohort: a day or wallet whose only trades
    # were excluded would otherwise get a row of zeros in every aggregate.
    day_dates = sorted(clean["date"].unique())
    wallets = sorted(clean["wallet_label"].unique())

    wb = Workbook()

    # ---------------------------------------------------------------- Raw Trades
    raw = wb.active
    raw.title = "Raw Trades"
    raw_headers = ["epoch_ms", "Date", "Wallet", "Mint", "Sell Signature", "Buy Count", "Hold (ms)",
                   "Wallet Token Amt", "Wallet Lamports Spent", "Wallet Lamports Received", "Wallet PnL (SOL)",
                   "Our Lamports Spent", "Our Lamports Received", "Our PnL (SOL)", "Wallet Win", "Our Win",
                   "Fill Ratio", "Blow-up", "Buy Reverted", "Sell Reverted", "Collector Ver"]
    raw.append(raw_headers)
    style_header_row(raw, 1, len(raw_headers))

    for i, row in df.iterrows():
        r = i + 2
        vals = [
            int(row["epoch_ms"]), row["date"], row["wallet_label"], row["mint"],
            row["sell_signature"] if pd.notna(row.get("sell_signature")) else "",
            int(row["buy_count"]), int(row["hold_duration_ms"]), float(row["wallet_token_amount"]),
            int(row["wallet_lamports_spent"]), int(row["wallet_lamports_received"]), float(row["wallet_pnl_sol"]),
            float(row["our_lamports_spent"]), float(row["our_lamports_received"]), float(row["our_pnl_sol"]),
            1 if row["wallet_pnl_sol"] > 0 else 0, 1 if row["our_pnl_sol"] > 0 else 0,
            float(row["fill_ratio"]) if pd.notna(row.get("fill_ratio")) else None,
            1 if bool(row["blowup"]) else 0,
            int(row["would_have_reverted"]), int(row["sell_would_have_reverted"]), int(row["collector_version"]),
        ]
        for c, v in enumerate(vals, start=1):
            cell = raw.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
        raw.cell(row=r, column=2).number_format = DATE_FMT
        for c in (9, 10):
            raw.cell(row=r, column=c).number_format = "#,##0"
        for c in (11, 12, 13, 14):
            raw.cell(row=r, column=c).number_format = SOL_FMT
        raw.cell(row=r, column=17).number_format = "0.000"

    autosize(raw, [14, 12, 12, 14, 20, 9, 10, 16, 18, 20, 13, 16, 18, 13, 10, 8, 11, 9, 12, 12, 12])
    raw.freeze_panes = "A2"

    # ---------------------------------------------------------- Daily Performance
    daily = wb.create_sheet("Daily Performance")
    daily.append(["Date", "Trades Closed", "Wallet Wins", "Wallet Win Rate", "Wallet PnL (SOL)",
                  "Our Wins", "Our Win Rate", "Our PnL (SOL)", "Cumulative Wallet PnL", "Cumulative Our PnL"])
    style_header_row(daily, 1, 10)

    cum_w = cum_o = 0.0
    for i, d in enumerate(day_dates):
        sub = clean[clean["date"] == d]
        nd = len(sub)
        ww = int((sub["wallet_pnl_sol"] > 0).sum())
        ow = int((sub["our_pnl_sol"] > 0).sum())
        wp = float(sub["wallet_pnl_sol"].sum())
        op = float(sub["our_pnl_sol"].sum())
        cum_w += wp
        cum_o += op
        r = i + 2
        for c, v in enumerate([d, nd, ww, ww / nd if nd else 0, wp, ow, ow / nd if nd else 0, op, cum_w, cum_o], 1):
            daily.cell(row=r, column=c, value=v).font = BODY_FONT
        daily.cell(row=r, column=1).number_format = DATE_FMT
        for c in (4, 7):
            daily.cell(row=r, column=c).number_format = PCT_FMT
        for c in (5, 8, 9, 10):
            daily.cell(row=r, column=c).number_format = SOL_FMT

    last_daily_row = len(day_dates) + 1
    autosize(daily, [12, 13, 12, 15, 15, 10, 14, 14, 19, 16])
    daily.freeze_panes = "A2"

    bar = BarChart()
    bar.type, bar.grouping = "col", "clustered"
    bar.title = "Daily P&L: Wallet (real) vs Our Simulated Copy"
    bar.y_axis.title, bar.x_axis.title = "SOL", "Date"
    bar.height, bar.width = 10, 24
    cats = Reference(daily, min_col=1, min_row=2, max_row=last_daily_row)
    bar.add_data(Reference(daily, min_col=5, max_col=5, min_row=1, max_row=last_daily_row), titles_from_data=True)
    bar.add_data(Reference(daily, min_col=8, max_col=8, min_row=1, max_row=last_daily_row), titles_from_data=True)
    bar.set_categories(cats)
    daily.add_chart(bar, "L2")

    line = LineChart()
    line.title = "Cumulative P&L Over Time"
    line.y_axis.title, line.x_axis.title = "SOL", "Date"
    line.height, line.width = 10, 24
    line.add_data(Reference(daily, min_col=9, max_col=10, min_row=1, max_row=last_daily_row), titles_from_data=True)
    line.set_categories(cats)
    daily.add_chart(line, "L22")

    # ------------------------------------------------------------------- By Wallet
    bywallet = wb.create_sheet("By Wallet")
    bywallet.append(["Wallet", "Trades Closed", "Wallet Win Rate", "Wallet PnL (SOL)", "Wallet Avg PnL/Trade",
                     "Our Win Rate", "Our PnL (SOL)", "Our Avg PnL/Trade"])
    style_header_row(bywallet, 1, 8)

    for i, wname in enumerate(wallets):
        sub = clean[clean["wallet_label"] == wname]
        nw = len(sub)
        wp = float(sub["wallet_pnl_sol"].sum())
        op = float(sub["our_pnl_sol"].sum())
        r = i + 2
        vals = [wname, nw, (sub["wallet_pnl_sol"] > 0).mean() if nw else 0, wp, wp / nw if nw else 0,
                (sub["our_pnl_sol"] > 0).mean() if nw else 0, op, op / nw if nw else 0]
        for c, v in enumerate(vals, 1):
            bywallet.cell(row=r, column=c, value=v).font = BODY_FONT
        for c in (3, 6):
            bywallet.cell(row=r, column=c).number_format = PCT_FMT
        for c in (4, 5, 7, 8):
            bywallet.cell(row=r, column=c).number_format = SOL_FMT

    last_bw_row = len(wallets) + 1
    autosize(bywallet, [14, 14, 15, 15, 18, 13, 14, 17])
    bywallet.freeze_panes = "A2"

    bw_chart = BarChart()
    bw_chart.type, bw_chart.grouping = "col", "clustered"
    bw_chart.title = "Total P&L by Wallet: Real vs Our Simulated Copy"
    bw_chart.y_axis.title = "SOL"
    bw_chart.height, bw_chart.width = 10, 24
    bw_chart.add_data(Reference(bywallet, min_col=4, max_col=4, min_row=1, max_row=last_bw_row), titles_from_data=True)
    bw_chart.add_data(Reference(bywallet, min_col=7, max_col=7, min_row=1, max_row=last_bw_row), titles_from_data=True)
    bw_chart.set_categories(Reference(bywallet, min_col=1, min_row=2, max_row=last_bw_row))
    bywallet.add_chart(bw_chart, "K2")

    # ---------------------------------------------------------------------- Summary
    s = wb.create_sheet("Summary", 0)
    put(s, "A1", "Paper Trading Performance Summary", TITLE_FONT)
    put(s, "A2", f"{n_clean} clean closed trades across {len(wallets)} tracked wallets "
                 f"({n_blow} blow-up, {n_buy_rev} reverting buy, {n_sell_rev} reverting sell held back).", NOTE_FONT)

    put(s, "A4", "Date range", LABEL_FONT)
    put(s, "B4", day_dates[0] if day_dates else None, BODY_FONT, DATE_FMT)
    put(s, "C4", day_dates[-1] if day_dates else None, BODY_FONT, DATE_FMT)

    put(s, "A5", "Total trades closed", LABEL_FONT)
    put(s, "B5", n_clean)
    put(s, "A6", "Days with activity", LABEL_FONT)
    put(s, "B6", len(day_dates))
    put(s, "A7", "Avg trades / active day", LABEL_FONT)
    put(s, "B7", n_clean / len(day_dates) if day_dates else 0, BODY_FONT, "0.0")

    put(s, "B9", "Wallet (real)", LABEL_FONT)
    put(s, "C9", "Our simulated copy", LABEL_FONT)

    cw_pnl = float(clean["wallet_pnl_sol"].sum())
    co_pnl = float(clean["our_pnl_sol"].sum())
    put(s, "A10", "Overall win rate", LABEL_FONT)
    put(s, "B10", (clean["wallet_pnl_sol"] > 0).mean() if n_clean else 0, BODY_FONT, PCT_FMT)
    put(s, "C10", (clean["our_pnl_sol"] > 0).mean() if n_clean else 0, BODY_FONT, PCT_FMT)
    put(s, "A11", "Total P&L (SOL)", LABEL_FONT)
    put(s, "B11", cw_pnl, BODY_FONT, SOL_FMT)
    put(s, "C11", co_pnl, BODY_FONT, SOL_FMT)
    put(s, "A12", "Avg P&L / trade (SOL)", LABEL_FONT)
    put(s, "B12", cw_pnl / n_clean if n_clean else 0, BODY_FONT, SOL_FMT)
    put(s, "C12", co_pnl / n_clean if n_clean else 0, BODY_FONT, SOL_FMT)
    put(s, "A13", "Avg P&L / active day (SOL)", LABEL_FONT)
    put(s, "B13", cw_pnl / len(day_dates) if day_dates else 0, BODY_FONT, SOL_FMT)
    put(s, "C13", co_pnl / len(day_dates) if day_dates else 0, BODY_FONT, SOL_FMT)

    put(s, "A15", "Data freshness", LABEL_FONT)
    put(s, "A16", "Newest trade (UTC)", LABEL_FONT)
    put(s, "B16", datetime.datetime.utcfromtimestamp(newest_epoch_ms / 1000), BODY_FONT, DATETIME_FMT)
    put(s, "A17", "Hours since newest trade (at generation)", LABEL_FONT)
    put(s, "B17", round(age_hours, 1), BODY_FONT, "0.0")
    put(s, "A18", "Status", LABEL_FONT)
    put(s, "B18", "STALE -- check the Actions tab" if age_hours > STALE_HOURS else "OK")
    put(s, "A19", "Freshness is measured when this file is generated, not when it is opened -- "
                  "regenerate to re-check.", SMALL_NOTE_FONT)

    put(s, "A20", "Note: 'Our simulated copy' assumes copying closed trades at the configured execution_lag_ms, "
                  "on top of the free poll engine's own ~17s median detection lag. It is NOT an unfiltered "
                  "population: reverting and blow-up fills are held out (sized below), and RPC rate-limiting on "
                  "the free tier drops an unmeasured share of the wallets' trades before they are ever seen.",
        NOTE_FONT)

    put(s, "A22", "Excluded cohorts (held out of every figure above)", LABEL_FONT)
    rows = [
        ("A23", f"Blow-up fills: cost outside {BLOWUP_LO}-{BLOWUP_HI}x the wallet's for the same tokens",
         n_blow, float(df.loc[df["blowup"], "our_pnl_sol"].sum())),
        ("A24", "Reverting BUYS: price passed the wallet's max_sol_cost bound during our lag",
         n_buy_rev, float(df.loc[df["would_have_reverted"] == 1, "our_pnl_sol"].sum())),
        ("A25", "Reverting SELLS: price fell through the wallet's min_sol_output bound during our lag",
         n_sell_rev, float(df.loc[df["sell_would_have_reverted"] == 1, "our_pnl_sol"].sum())),
    ]
    for cell, label, cnt, pnl in rows:
        r = int(cell[1:])
        put(s, cell, label)
        put(s, f"B{r}", cnt)
        put(s, f"C{r}", pnl, BODY_FONT, SOL_FMT)

    put(s, "A26", "Our total P&L INCLUDING all of them (SOL)", LABEL_FONT)
    # Summed off the full frame, not by adding the cohorts: a row can be both a
    # blow-up and a reverting fill, and adding them would double-count it.
    put(s, "B26", float(df["our_pnl_sol"].sum()), BODY_FONT, SOL_FMT)

    put(s, "A28", "Collector version (which round trips got written at all)", LABEL_FONT)
    ver_notes = {
        1: "v1 -- CENSORED: dropped reverting buys AND reverting sells outright",
        2: "v2 -- records reverting buys; still dropped round trips with a reverting sell",
        3: "v3 -- complete: records both",
    }
    r = 29
    for ver in sorted(df["collector_version"].unique()):
        put(s, f"A{r}", ver_notes.get(int(ver), f"v{int(ver)}"))
        put(s, f"B{r}", int((df["collector_version"] == ver).sum()))
        r += 1
    put(s, f"A{r}", "Only the newest collector version gives an unbiased picture. Rows from earlier versions are "
                    "missing trades entirely -- those were never written and cannot be recovered, so their totals "
                    "read better than reality rather than merely noisier.", SMALL_NOTE_FONT)

    autosize(s, [62, 16, 20])
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}: {n} trades ({n_clean} clean), {len(day_dates)} days, {len(wallets)} wallets")


if __name__ == "__main__":
    main()
